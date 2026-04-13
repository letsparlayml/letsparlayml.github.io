#!/usr/bin/env python3
import argparse
import csv
import json
from pathlib import Path
from datetime import datetime, UTC
import re

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

STATUS_MAP = {
    "p": "Probable",
    "probable": "Probable",
    "q": "Questionable",
    "questionable": "Questionable",
    "gtd": "Questionable",
    "out": "Out",
    "o": "Out",
    "d": "Doubtful",
    "doubtful": "Doubtful",
}

def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")

def canonical_status(value: str) -> str:
    key = str(value or "").strip().lower()
    return STATUS_MAP.get(key, str(value or "").strip().title())

def read_rows_from_xlsx(path: Path):
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read the Excel template.")
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = wb["injuries_input"] if "injuries_input" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_header(h) for h in rows[0]]
    output = []
    for raw in rows[1:]:
        if not raw or all(v in (None, "") for v in raw):
            continue
        row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
        output.append(row)
    return output

def read_rows_from_csv(path: Path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            rows.append({normalize_header(k): v for k, v in row.items()})
        return rows

def convert_value(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            return str(value)
    return str(value).strip()

def pick_default_input(repo: Path) -> Path:
    candidates = [
        repo / "data" / "nba_injuries.xlsx",
        repo / "data" / "nba_injuries.csv",
        repo / "nba_injuries.xlsx",
        repo / "nba_injuries.csv",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return repo / "data" / "nba_injuries.xlsx"

def main():
    parser = argparse.ArgumentParser(description="Build nba_injuries.json from a manual Excel or CSV file.")
    parser.add_argument("--website-repo", required=True, help="Path to the website repo root.")
    parser.add_argument("--input", help="Optional explicit path to the Excel or CSV injury file.")
    parser.add_argument("--output", help="Optional explicit output path for nba_injuries.json.")
    args = parser.parse_args()

    repo = Path(args.website_repo)
    input_path = Path(args.input) if args.input else pick_default_input(repo)
    output_path = Path(args.output) if args.output else repo / "data" / "nba_injuries.json"

    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    if input_path.suffix.lower() == ".csv":
        rows = read_rows_from_csv(input_path)
    else:
        rows = read_rows_from_xlsx(input_path)

    players = []
    for row in rows:
        player = convert_value(row.get("player_name") or row.get("player"))
        status = canonical_status(convert_value(row.get("status")))
        if not player or not status:
            continue
        players.append({
            "gameDate": convert_value(row.get("game_date")),
            "player": player,
            "team": convert_value(row.get("team") or row.get("team_abbr")),
            "status": status,
            "note": convert_value(row.get("injury_note") or row.get("note")),
            "source": convert_value(row.get("source")),
            "lastUpdated": convert_value(row.get("last_updated") or row.get("updated_at")),
        })

    players.sort(key=lambda x: (x["gameDate"], x["team"], x["player"]))

    payload = {
        "updatedAt": datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "sourceFile": input_path.name,
        "players": players,
    }

    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {len(players)} injury rows to {output_path}")
    print(f"Source file: {input_path}")

if __name__ == "__main__":
    main()
